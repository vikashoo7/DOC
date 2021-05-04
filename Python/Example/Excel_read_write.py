from openpyxl import Workbook
from openpyxl.styles import Font

wb = Workbook()		###declear the workbook
ws = wb.active		###initialising the workbook


###initialise the workbook sheet. summary -> name of worksheet. 0->worksheet number
smry = wb.create_sheet('Summary',0)	

###Declearing the colum name
smry['A1'] = "Server Name"		
smry['B1'] = "user name"

###providing the shell size
smry.column_dimensions['A'].width = 40	
smry.column_dimensions['B'].width = 15	

###make the heading in BOLD
smry.cell(row = 1, column = 1).font = Font( bold = True)
smry.cell(row = 1, column = 2).font = Font( bold = True)


###initialise the another workbook sheet. "ALL INVENTORY" -> name of worksheet. 1->worksheet number ### we can add many worksheet as per the out need
all = wb.create_sheet('ALL INVENTORY',1)

###Declearing the colument name
all['A1'] = "FQDN"
all['B1'] = "IP Address"

###make the heading in BOLD
all['A1'].font = Font( bold = True)
all['B1'].font = Font( bold = True)

###providing the shell size
all.column_dimensions['A'].width = 40
all.column_dimensions['B'].width = 15



###Removing the empty sheet which will create by default
empty_sheet=wb.get_sheet_by_name('Sheet')
wb.remove_sheet(empty_sheet)

###link to the value from one to another.
k=1
app_name='ALL INVENTORY'
link = '=HYPERLINK("#' +"'" + app_name[0:31] +"'"+ '!A1", "' + app_name + '")'

SA = 'A' +str(k + 1)
smry[SA] = link


####Save the Excel Sheet
wb.save('/home/user/inventory.xlsx')

